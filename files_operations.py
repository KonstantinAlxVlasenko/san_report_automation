import sys
import os
import openpyxl
import re
import json
import xlrd
import pandas as pd
from os import makedirs
from datetime import date

'''Module to perform operations with files (create folder, save data to excel file, import data from excel file)'''


def create_folder(path, max_title):
    """Function to create any folder with 'path' 
    """    
    info = f'Make directory {os.path.basename(path)}'
    print(info, end = ' ')
    # if folder not exist create 
    if not os.path.exists(path):        
        try:
            os.makedirs(path)
        except OSError:
            status_info('fail', max_title, len(info))
            print(f'Not possible create directory {path}.')
            print('Code execution finished')
            sys.exit()
        else:
            status_info('ok', max_title, len(info))
    # otherwise print 'SKIP' status
    else:
        status_info('skip', max_title, len(info))
        
        
def check_valid_path(path):
    """Function to check if folder exist
    """
    path = os.path.normpath(path)
    if not os.path.isdir(path):
        print(f"{path} folder doesn't exist")
        print('Code execution exit')
        sys.exit()
        
        
# saving DataFrame to excel file
def save_xlsx_file(data_frame, sheet_title, report_data_lst, report_type = 'service'):   
    """Check if excel file exists, check if dataframe sheet is in file, 
    delete sheet with stored dataframe (if sheet tabs number > 1)
    and save new dataframe 
    """    
    customer_name, report_path, _, max_title, report_steps_dct = report_data_lst      
    current_date = str(date.today())
    # construct excel filename
    file_name = customer_name + '_' + report_type + '_' + 'report_' + current_date + '.xlsx'
    # information string
    info = f'Exporting {sheet_title} to {report_type} file'
    print(info, end =" ")
    file_path = os.path.join(report_path, file_name)
    
    # save DataFrame to excel file if export_to_excel trigger is ON
    # and DataFrame is not empty
    if report_steps_dct[sheet_title][0] and not data_frame.empty:
        # pd.ExcelWriter has only two file open modes
        # if file doesn't exist it has be opened in "w" mode otherwise in "a"    
        if os.path.isfile(file_path):
            file_mode = 'a'
            # open existing excel file
            workbook = openpyxl.load_workbook(file_path)
            if sheet_title in workbook.sheetnames:
                # # import stored data in excel file do dataframe
                # import_df = pd.read_excel(file_path, sheet_name=sheet_title)
                # # check if new dataframe and stored in excelfile are equal
                # if not data_frame.equals(import_df):
                #     # if dataframes are not equal delete sheet
                
                # after sheet removal excel file must contain at least one sheet
                if len(workbook.sheetnames) != 1:                    
                    del workbook[sheet_title]
                    try:
                        workbook.save(file_path)
                    except PermissionError:
                        status_info('fail', max_title, len(info))
                        print('Permission denied. Close the file.')
                        sys.exit()
                else:
                    file_mode = 'w'
                    # export new dataframe
                    # export_dataframe(file_path, data_frame, sheet_title, file_mode,str_length_status)
                # # if dataframes are equal then skip
                # else:
                #     file_mode = None
                #     status_info('skip', max_title, len(info))                  
        else:
            file_mode = 'w'
        
        # if required export DataFrame to the new sheet
        if file_mode:
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl',  mode=file_mode) as writer:
                    # check if DataFrame have MultiIndex
                    # reset index if True
                    if isinstance(data_frame.index, pd.MultiIndex):
                        data_frame_flat = data_frame.reset_index()
                    # keep index if False
                    else:
                        data_frame_flat = data_frame.copy()
                    # saving DataFrame with single Index
                    data_frame_flat.to_excel(writer, sheet_name=sheet_title, index=False)
            except PermissionError:
                status_info('fail', max_title, len(info))
                print('Permission denied. Close the file.')
                sys.exit()
            else:
                status_info('ok', max_title, len(info))        
    else:
        # if save key is on but DataFrame empty
        if report_steps_dct[sheet_title][0] and data_frame.empty:
            status_info('no data', max_title, len(info))
        else:            
            status_info('skip', max_title, len(info))

def status_info(status, max_title, len_info_string, shift=0):
    """Function to print current operation status ('OK', 'SKIP', 'FAIL')
    """    
    # information + operation status string length in terminal
    str_length = max_title + 80 + shift
    status = status.upper()
    # status info aligned to the right side
    # space between current operation information and status of its execution filled with dots
    print(status.rjust(str_length - len_info_string, '.'))

 
def columns_import(sheet_title, max_title, *args, init_file = 'san_automation_info.xlsx'):
    """Function to import corresponding columns from init file.
    Can import several columns.
    """
    # file to store all required data to process configuratin files
    # default init_file  is 'san_automation_info.xlsx'
    # columns titles string to display imported columns list without parenthesis
    columns_str = ""
    for arg in args:
        columns_str += "'" + arg + "', "
    columns_str = columns_str.rstrip(", ")

    info = f'Importing {columns_str} from {sheet_title} tab'
    print(info, end = ' ')
    # try read data in excel
    try:
        columns = pd.read_excel(init_file, sheet_name = sheet_title, usecols = args, squeeze=True)
    except FileNotFoundError:
        status_info('fail', max_title, len(info))
        print(f'File not found. Check if file {init_file} exist.')
        sys.exit()
    except ValueError:
        status_info('fail', max_title, len(info))
        print(f'Column(s) {columns_str} not found. Check if column exist in {sheet_title}.')
        sys.exit()        
    else:
        # if number of columns to read > 1 than returns corresponding number of lists
        if len(args)>1:
            columns_names = [columns[arg].dropna().tolist() if not columns[arg].empty else None for arg in args]
        else:
            columns_names = columns.dropna().tolist()
        status_info('ok', max_title, len(info))
    
    return columns_names


def dataframe_import(sheet_title, max_title, init_file = 'san_automation_info.xlsx'):
    """Function to import dataframe from exel file"""
    # file to store all required data to process configuratin files
    # init_file = 'san_automation_info.xlsx'   
    info = f'Importing {sheet_title} dataframe from {init_file} file'
    print(info, end = ' ')
    # try read data in excel
    try:
        dataframe = pd.read_excel(init_file, sheet_name = sheet_title)
    # if file is not found
    except FileNotFoundError:
        status_info('fail', max_title, len(info))
        print(f'File not found. Check if file {init_file} exists.')
        sys.exit()
    # if sheet is not found
    except xlrd.biffh.XLRDError:
        status_info('fail', max_title, len(info))
        print(f'Sheet {sheet_title} not found in {init_file}. Check if it exists.')
        sys.exit()
    else:
        status_info('ok', max_title, len(info))
    
    return dataframe


def data_extract_objects(sheet_title, max_title):
    """Function imports parameters names and regex tepmplates
    to extract required data from configuration files   
    """
    # imports keys to extract switch parameters from tmp dictionary
    params_names, params_add_names = columns_import(sheet_title, max_title, 'params', 'params_add')
    # imports base names for compile and match templates and creates corresonding names
    keys = columns_import(sheet_title, max_title, 're_names')
    comp_keys = [key+'_comp' for key in keys]
    match_keys = [key + '_match' for key in keys]
    # imports string for regular expressions
    comp_values = columns_import(sheet_title,  max_title, 'comp_values')
    # creates regular expressions
    comp_values_re = [re.compile(fr"{element}") for element in comp_values]
    # creates dictionary with regular expressions  
    comp_dct = dict(zip(comp_keys, comp_values_re))
    
    return params_names, params_add_names, comp_keys, match_keys, comp_dct


def export_lst_to_excel(data_lst, report_data_lst, sheet_title_export, sheet_title_import = None, 
                        columns = columns_import, columns_title_import = 'columns'):
    """Function to export list to DataFrame and then save it to excel report file
    returns DataFrame
    """    
    *_, max_title, _ = report_data_lst 
    
    # checks if columns were passed to function as a list
    if isinstance(columns, list):
        columns_title = columns
    # if not (default) then import columns from excel file
    else:
        columns_title =columns(sheet_title_import, max_title, columns_title_import)
    data_df = pd.DataFrame(data_lst, columns= columns_title)
    save_xlsx_file(data_df, sheet_title_export, report_data_lst)
    
    return data_df


def line_to_list(re_object, line, *args):
    """Function to extract values from line with regex object 
    and combine values with other optional data into list
    """
    values, = re_object.findall(line)
    if isinstance(values, tuple) or isinstance(values, list):
        values_lst = [value.rstrip() if value else None for value in values]
    else:
        values_lst = [values.rstrip()]
    return [*args, *values_lst]


def update_dct(keys, values, dct, char = ', '):
    """Function to add param_add:value pairs
    to the dictionary with discovered parameters
    """
    for key, value in zip(keys, values):
        if value:                
            if isinstance(value, set) or isinstance(value, list):
                value = f'{char}'.join(value)
            dct[key] = value
    return dct


def dct_from_columns(sheet_title, max_title, *args, init_file = 'report_info.xlsx'):
    """Function imports columns and create dictionary
    If only one column imported then dictionary with keys and empty lists as values created
    If several columns imported then first column is keys of dictionary and others are values
    """
    # info string in case if not possible to create dictionary
    info = f'{args} columns have different length. Not able to create dictionary. Check data in {sheet_title} tab'

    # if one column is passed then create dictionary with keys and empty lists as values for each key
    if len(args) == 1:
        keys = columns_import(sheet_title, max_title, *args, init_file)
        dct = dict((key, []) for key in keys)
    # if two columns passed then create dictionary of keys with one value for each key
    elif len(args) == 2:
        keys, values = columns_import(sheet_title, max_title, *args, init_file = init_file)
        # if columns have different number of elements throw information string and exit
        if len(keys) != len(values):
            print(info)
            sys.exit()                    
        dct ={key: value for key, value in zip(keys, values)}
    # if morte than two columns passed then create dictionary of keys with list of values for each key
    elif len(args) > 2:
        # first column is keys rest columns are in values list of lists
        keys, *values = columns_import(sheet_title, max_title, *args, init_file = init_file)
        # check if all imported columns have equal length to create dictionary
        # create set of columns length with set comprehension method
        columns_len_set = {len(columns_title) for columns_title in [keys, *values]}
        # columns length set contains more than 1 element show information string
        if len(columns_len_set) != 1:
            print(info)
            sys.exit()
        # dictionary with key and value as list of lists 
        dct ={key: value for key, *value in zip(keys, *values)}
    else:
        print('Not sufficient data to create dictionary')
        sys.exit()

    return dct

def force_extract_check(data_names, data_lst, force_extract_keys_lst, max_title):
    """Function to check if force data extract key is ON
    If all data from data list are present and True(non-zero) but extract key is ON
    Then function prints for which data force extraction has been initiated
    Returns list with data check results 
    """
    # list to store data check results
    data_check = []
    # check each data in data_lst
    for data in data_lst:
        # for DataFrame method empty is used to check if DataFrame has data
        if isinstance(data, pd.DataFrame):
            data_check.append(not data.empty)
        # for bool type append True (means value is present) no matter if value is True or False
        elif isinstance(data, bool):
            data_check.append(True)
        # for other types of data no special method needed
        else:
            data_check.append(data)

    # when all data are not empty but force exctract is ON
    # print data names for which extraction is forced
    if all(data_check) and any(force_extract_keys_lst):
        # list for which data extraction is forced
        force_extract_names_lst = [data_name for data_name, force_extract_key in zip(data_names, force_extract_keys_lst) if force_extract_key]
        info = f'Force {", ".join(force_extract_names_lst)} data extract initialize'
        print(info, end =" ")
        status_info('ok', max_title, len(info))
        
    return data_check
            

def save_data(report_data_list, data_names, *args):
    """Function to export extracted configuration data to JSON or CSV file
    depending on data passed 
    """   
    customer_name, _, json_data_dir, max_title, _ = report_data_list    
    # data_names it's a list of names for data passed as args
    empty_data = None

    for data_name, data_exported in zip(data_names, args):
        file_name = customer_name + '_' + data_name
        # adding file extenson depending from type of data saved
        # csv for DataFrame
        if isinstance(data_exported, pd.DataFrame):
            file_name += '.csv'
        # for all other types is json
        else:
            # create file name and path for json file
            file_name += '.json'
        # full file path
        file_path = os.path.join(json_data_dir, file_name)
        info = f'Saving {data_name} to {file_name} file'
        try:
            print(info, end =" ")
            with open(file_path, 'w', encoding="utf-8") as file:
                # savng data for DataFrame
                if isinstance(data_exported, pd.DataFrame):
                    if not data_exported.empty:
                        # check if DataFrame have MultiIndex
                        # reset index if True due to MultiIndex is not saved                    
                        if isinstance(data_exported.index, pd.MultiIndex):
                            data_exported_flat = data_exported.reset_index()
                        # keep indexing if False
                        else:
                            data_exported_flat = data_exported.copy()
                        # save single level Index DataFrame to csv
                        data_exported_flat.to_csv(file, index=False)
                    else:
                        empty_data = True
                # for all other types
                else:
                    if len(data_exported) != 0: 
                        json.dump(data_exported, file)
                    else:
                        empty_data = True
        # display writing data to file status
        except FileNotFoundError:
            status_info('fail', max_title, len(info))
        else:
            if not empty_data:
                status_info('ok', max_title, len(info))
            else:
                status_info('empty', max_title, len(info))
                
                        
def load_data(report_data_list, *args):
    """Function to load data from JSON or CSV file to data object
    Detects wich type of data required to be loaded automaticaly
    Returns list with imported data 
    """    
    customer_name, _, json_data_dir, max_title, _ = report_data_list
    # list to store loaded data
    data_imported = []
    
    # check real file path
    for data_name in args:
        # constructing filenames for json and csv files
        file_name_json = customer_name + '_' + data_name + '.json'
        file_path_json = os.path.join(json_data_dir, file_name_json)
        file_name_csv = customer_name + '_' + data_name + '.csv'
        file_path_csv = os.path.join(json_data_dir, file_name_csv)
        # flags file existance
        file_json = False
        file_csv = False
        # check if json file exist
        if os.path.isfile(file_path_json):
            file_path = file_path_json
            file_name = file_name_json
            file_json = True
        # check if csv file exist
        elif os.path.isfile(file_path_csv):
            file_name = file_name_csv
            file_path = file_path_csv
            file_csv = True
        # if no files exist then dislay info about no data availabilty
        # and add None to data_imported list
        else:
            info = f'Loading {data_name}'
            print(info, end =" ")
            status_info('no data', max_title, len(info))
            data_imported.append(None)
        
        # when saved file founded 
        if any([file_json, file_csv]):                   
            info = f'Loading {data_name} from {file_name} file'
            print(info, end =" ")
            # open file
            try:  
                with open(file_path, 'r', encoding="utf-8") as file:
                    # for json file use load method
                    if file_json:
                        data_imported.append(json.load(file))
                    # for csv file use read_csv method
                    elif file_csv:
                        data_imported.append(pd.read_csv(file))
            # display file load status
            except:
                status_info('no data', max_title, len(info))
                data_imported.append(None)
            else:
                status_info('ok', max_title, len(info))
           
    return data_imported             