'''Module to perform operations with files (create folder, save data to excel file, import data from excel file)'''

# import re
import json
import os
import sys
# from os import makedirs
from datetime import date

import openpyxl
# import xlrd
import pandas as pd

from common_operations_miscellaneous import status_info


def create_folder(path, max_title):
    """Function to create any folder with path"""

    info = f'Make directory {os.path.basename(path)}'
    print(info, end = ' ')
    # if folder not exist create 
    if not os.path.exists(path):        
        try:
            os.makedirs(path)
        except OSError:
            status_info('fail', max_title, len(info))
            print(f'Not possible create directory {path}.')
            print('Code execution finished')
            sys.exit()
        else:
            status_info('ok', max_title, len(info))
    # otherwise print 'SKIP' status
    else:
        status_info('skip', max_title, len(info))
        
     
def check_valid_path(path):
    """Function to check if folder exist"""

    path = os.path.normpath(path)
    if not os.path.isdir(path):
        print(f"{path} folder doesn't exist")
        print('Code execution exit')
        sys.exit()
        

# saving DataFrame to excel file
def save_xlsx_file(data_frame, sheet_title, report_data_lst, force_flag = False):
    """
    Check if excel file exists, check if dataframe sheet is in file, 
    delete sheet with stored dataframe (if sheet tabs number > 1)
    and save new dataframe 
    """
    
    customer_name, report_path, _, max_title, report_steps_dct = report_data_lst

    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False
    
    # report_steps_dct for each data_name contains: export_to_excel flag, 
    # force_extract flag, report_type, step_info, data_description
    if report_steps_dct.get(sheet_title):
        export_flag, _, report_type, _, df_decription = report_steps_dct.get(sheet_title)
    else:
        info = f'DataFrame {sheet_title}'
        print(info, end =" ")
        status_info('unknown', max_title, len(info))
        export_flag, report_type = 1, 'unknown'
        df_decription = '-'

    # check DataFrame report type to save
    if report_type == 'report':
        report_mark = 'SAN_Assessment_Tables'
    else:
        report_mark = report_type
          
    current_date = str(date.today())
    # construct excel filename
    file_name = customer_name + '_' + report_mark + '_' + current_date + '.xlsx'

    # information string
    info = f'Exporting {sheet_title} table to {report_mark} file'
    print(info, end =" ")
    file_path = os.path.join(report_path, file_name)
    
    # save DataFrame to excel file if export_to_excel trigger is ON
    # and DataFrame is not empty
    if (force_flag or export_flag) and not data_frame.empty:
        # pd.ExcelWriter has only two file open modes
        # if file doesn't exist it has be opened in "w" mode otherwise in "a"    
        if os.path.isfile(file_path):
            file_mode = 'a'
            # open existing excel file
            workbook = openpyxl.load_workbook(file_path)
            if sheet_title in workbook.sheetnames:
                # after sheet removal excel file must contain at least one sheet
                if len(workbook.sheetnames) != 1:                    
                    del workbook[sheet_title]
                    try:
                        workbook.save(file_path)
                    except PermissionError:
                        status_info('fail', max_title, len(info))
                        print('\nPermission denied. Close the file.\n')
                        sys.exit()
                else:
                    file_mode = 'w'                
        else:
            file_mode = 'w'
        
        # if required export DataFrame to the new sheet
        if file_mode:
            try:
                with pd.ExcelWriter(file_path, mode=file_mode, options=options) as writer:  # pylint: disable=abstract-class-instantiated
                    # table of content item generation start
                    # if file is new then create table of content
                    if file_mode == 'w':
                        # save current item with header
                        content_df = pd.DataFrame([[sheet_title, df_decription]], columns=['Закладка', 'Название таблицы'])
                        content_df.to_excel(writer, sheet_name='Содержание', index=False)
                        writer.save()
                        # create hyperlink to wotksheet
                        workbook = openpyxl.load_workbook(file_path)
                        writer.book = workbook
                        writer.sheets = {ws.title: ws for ws in workbook.worksheets}
                        startrow = workbook['Содержание'].max_row
                        cell_address = 'A' + str(startrow)
                        create_hyperlink(workbook['Содержание'], cell_address, sheet_name = sheet_title, cell_ref='A1', display_name=None)
                    # if file exist new table of content item is appended if it's not exist
                    elif  file_mode == 'a':
                        workbook = openpyxl.load_workbook(file_path)
                        # check if item exist in menu
                        item_exist = []                        
                        for row in workbook.get_sheet_by_name('Содержание').values:
                            if any(value == sheet_title for value in row):
                                item_exist.append(True)
                            else:
                                item_exist.append(False)
                        # if not exist add item without header and create hyperlink
                        if not any(item_exist):
                            # it makes append item to the same worksheet instead of making new one
                            writer.book = workbook
                            writer.sheets = {ws.title: ws for ws in workbook.worksheets}
                            # content item to add
                            content_df = pd.DataFrame([[sheet_title, df_decription]], columns=['Закладка', 'Название таблицы'])
                            # find row and cell to add new item and hyperlink
                            startrow = writer.sheets['Содержание'].max_row
                            cell_address = 'A' + str(startrow+1)
                            # add content item
                            content_df.to_excel(writer, sheet_name='Содержание', startrow=startrow, index = False,header= False)
                            # create hyperlink
                            create_hyperlink(workbook['Содержание'], cell_address, sheet_name = sheet_title, cell_ref='A1', display_name=None)
                    # table of content item generation finish
                    # DataFrame save to worksheet start
                    # check if DataFrame have MultiIndex
                    # reset index if True
                    if isinstance(data_frame.index, pd.MultiIndex):
                        data_frame_flat = data_frame.reset_index()
                    # keep index if False
                    else:
                        data_frame_flat = data_frame.copy()
                    # saving DataFrame with single Index
                    data_frame_flat.to_excel(writer, sheet_name=sheet_title,  startrow=2, index=False)
                    writer.save()
                    workbook = openpyxl.load_workbook(file_path)
                    writer.book = workbook
                    writer.sheets = {ws.title: ws for ws in workbook.worksheets}
                    # add table title
                    workbook[sheet_title]['A1'] = df_decription
                    workbook[sheet_title]['A1'].font =  openpyxl.styles.fonts.Font(bold=True)
                    # add hyperlink to the Table of contents
                    create_hyperlink(workbook[sheet_title], 'A2', 'Содержание', cell_ref='A2', display_name='К содержанию')
                    workbook.save(file_path)
                    # DataFrame save to worksheet finish
            except PermissionError:
                status_info('fail', max_title, len(info))
                print('\nPermission denied. Close the file.\n')
                sys.exit()
            else:
                status_info('ok', max_title, len(info))
        return file_path        
    else:
        # if save key is on but DataFrame empty
        if report_steps_dct[sheet_title][0] and data_frame.empty:
            status_info('no data', max_title, len(info))
        else:            
            status_info('skip', max_title, len(info))
        return None


def save_data(report_data_list, data_names, *args):
    """
    Function to export extracted configuration data to JSON or CSV file
    depending on data passed 
    """

    customer_name, _, json_data_dir, max_title, _ = report_data_list    
    # data_names it's a list of names for data passed as args

    for data_name, data_exported in zip(data_names, args):
        empty_data = False
        file_name = customer_name + '_' + data_name
        # adding file extenson depending from type of data saved
        # csv for DataFrame
        if isinstance(data_exported, pd.DataFrame):
            file_name += '.csv'
        # for all other types is json
        else:
            # create file name and path for json file
            file_name += '.json'
        # full file path
        file_path = os.path.join(json_data_dir, file_name)
        info = f'Saving {data_name} to {file_name} file'
        try:
            print(info, end =" ")
            with open(file_path, 'w', encoding="utf-8") as file:
                # savng data for DataFrame
                if isinstance(data_exported, pd.DataFrame):
                    # check if DataFrame have MultiIndex
                    # reset index if True due to MultiIndex is not saved                    
                    if isinstance(data_exported.index, pd.MultiIndex):
                        data_exported_flat = data_exported.reset_index()
                    # keep indexing if False
                    else:
                        data_exported_flat = data_exported.copy()
                    # when DataFrame is empty fill first row values
                    # with information string
                    if data_exported_flat.empty:
                        empty_data = True
                        data_exported_flat.loc[0] = 'NO DATA FOUND'
                    # save single level Index DataFrame to csv
                    data_exported_flat.to_csv(file, index=False)

                # for all other types
                else:
                    # if data list is not empty
                    if not len(data_exported):
                        empty_data = True 
                        data_exported = 'NO DATA FOUND'
                    json.dump(data_exported, file)
        # display writing data to file status
        except FileNotFoundError:
            status_info('fail', max_title, len(info))
        else:
            if not empty_data:
                status_info('ok', max_title, len(info))
            else:
                status_info('empty', max_title, len(info))
                

def load_data(report_data_list, *args):
    """Function to load data from JSON or CSV file to data object
    Detects wich type of data required to be loaded automaticaly
    Returns list with imported data 
    """
        
    customer_name, _, json_data_dir, max_title, _ = report_data_list
    # list to store loaded data
    data_imported = []
    
    # check real file path
    for data_name in args:
        # constructing filenames for json and csv files
        file_name_json = customer_name + '_' + data_name + '.json'
        file_path_json = os.path.join(json_data_dir, file_name_json)
        file_name_csv = customer_name + '_' + data_name + '.csv'
        file_path_csv = os.path.join(json_data_dir, file_name_csv)
        # flags file existance
        file_json = False
        file_csv = False
        # check if json file exist
        if os.path.isfile(file_path_json):
            file_path = file_path_json
            file_name = file_name_json
            file_json = True
        # check if csv file exist
        elif os.path.isfile(file_path_csv):
            file_name = file_name_csv
            file_path = file_path_csv
            file_csv = True
        # if no files exist then dislay info about no data availabilty
        # and add None to data_imported list
        else:
            info = f'Loading {data_name}'
            print(info, end =" ")
            status_info('no file', max_title, len(info))
            data_imported.append(None)
        
        # when saved file founded 
        if any([file_json, file_csv]):                   
            info = f'Loading {data_name} from {file_name} file'
            print(info, end =" ")
            # open file
            try:  
                with open(file_path, 'r', encoding="utf-8") as file:
                    # for json file use load method
                    if file_json:
                        data_imported.append(json.load(file))
                    # for csv file use read_csv method
                    elif file_csv:
                        data_imported.append(pd.read_csv(file, dtype='unicode'))
            # display file load status
            except:
                status_info('no data', max_title, len(info))
                data_imported.append(None)
            else:
                status_info('ok', max_title, len(info))
           
    return data_imported


def create_files_list(folder, file_extension, max_title,):
    """
    Function to create list with files. Takes directory to check
    and file extension as a parameters. Returns list of files with the extension
    founded in root folder and nested folders.
    """
    info = f'Checking {os.path.basename(folder)} for configuration files'
    print(info, end =" ") 

    # check if ssave_path folder exist
    check_valid_path(folder)
   
    # list to save configuration data files
    files_lst = []

    # going through all directories inside ssave folder to find configurutaion data
    for root, _, files in os.walk(folder):
        for file in files:
            if file.endswith(file_extension):
                file_path = os.path.normpath(os.path.join(root, file))
                files_lst.append(file_path)

    if len(files_lst) == 0:
        status_info('no data', max_title, len(info))
    else:
        status_info('ok', max_title, len(info))
              
    return files_lst


def create_hyperlink(ws, at_cell, sheet_name, cell_ref='A1', display_name=None):
    if display_name is None:
        display_name = sheet_name
    to_location = "'{0}'!{1}".format(sheet_name, cell_ref)
    ws[at_cell].hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(display=display_name, ref=at_cell, location=to_location)
    ws[at_cell].value = display_name
    ws[at_cell].font =  openpyxl.styles.fonts.Font(u='single', color=openpyxl.styles.colors.BLUE)             
