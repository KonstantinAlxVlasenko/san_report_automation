import re

import pandas as pd
import utilities.dataframe_operations as dfop
from openpyxl import load_workbook


def interconnect_module_extract(synergy_config):
    """Function to extract enclosure and interconnect modules information 
    (enclosures and interconnectbays tabs of the meddler file"""

    syn_enclosure_df = pd.read_excel(synergy_config, sheet_name='enclosures')
    syn_module_df = pd.read_excel(synergy_config, sheet_name='interconnectbays')

    # enclosure information
    enclosure_columns = ['name', 'enclosuremodel', 'version', 'serialnumber']
    synergy_enclosure_df = syn_enclosure_df[enclosure_columns].copy()

    synergy_enclosure_df['enclosuretype'] = \
        synergy_enclosure_df[['enclosuremodel', 'version']].apply(dfop.wise_combine, axis=1)

    synergy_enclosure_df.rename(columns={'name': 'enclosurename', 'serialnumber': 'enclosure_serialnumber'}, inplace=True)
    synergy_enclosure_df.drop(columns=['enclosuremodel', 'version'], inplace=True)

    # interconnect information
    module_columns = ['enclosurename',  'baynumber',  'interconnectmodel',  'serialnumber',  'switchfwversion',  'hostname',  'switchbasewwn']
    mask_state = syn_module_df['state'].notna()
    synergy_module_df = syn_module_df.loc[mask_state, module_columns]

    synergy_module_df['device_location'] = \
        synergy_module_df[['enclosurename', 'baynumber']].apply(dfop.wise_combine, axis=1, args=('Enclosure ', ' bay '))

    # add interconnect information to enclosure
    synergy_module_df = synergy_enclosure_df.merge(synergy_module_df, how='left', on=['enclosurename'])
    return synergy_module_df


def server_mezz_extract(synergy_config, pattern_dct):
    """Function to extract synergy servers and servers mezzanine information"""
    
    # server and mezzanine information
    synergy_server_wwn_df = synergy_server_wwn(synergy_config, pattern_dct)
    # add mezzanin information to synergy_server_wwn_df from 'server-prof-conn-details' if sheet exist
    synergy_profile_wwn_df = synergy_profile_wwn(synergy_config, synergy_server_wwn_df, pattern_dct)
    # conctenate connection profile and server hardware
    synergy_servers_df = pd.concat([synergy_server_wwn_df, synergy_profile_wwn_df], ignore_index=True)
    synergy_servers_df.drop_duplicates(inplace=True)
    # add mezzanine firmware details
    synergy_servers_df = synergy_mezz_fw(synergy_config, synergy_servers_df, pattern_dct)
    synergy_servers_df.sort_values(by=['enclosurename', 'position', 'Mezz_WWPN'], ignore_index=True, inplace=True)
    return synergy_servers_df


def synergy_server_wwn(synergy_config, pattern_dct):
    """Server and mezzanine information from server-hardware tab"""

    syn_server_hw_df = pd.read_excel(synergy_config, sheet_name='server-hardware')
    
    # server_hardware
    server_hw_columns = [ 'enclosurename', 'position', 'servername',  'name',  'serverprofilename',  'model',  'serialnumber',  'oshint']
    # columns with Mezz in name except MAC columns
    server_mezz_columns = [column for column in syn_server_hw_df.columns if
                           re.match(pattern_dct['mezz_column'], column)]
    
    server_hw_mezz_columns = [*server_hw_columns, *server_mezz_columns]
    # filter servers (filter off composer)
    mask_profile = syn_server_hw_df['serverprofilename'] != 'None'
    # create DataFrame with server mezzanin cards
    synergy_mezz_df = syn_server_hw_df.loc[mask_profile, server_hw_mezz_columns].copy()


    # filter off empty mezzanin slots and columns except Mezz
    mezz_columns = [column for column in server_mezz_columns if
                    re.match(pattern_dct['mezz_number_column'], column) and
                    (synergy_mezz_df[column] != '<empty>').all()]

    # create synergy_server_wwn_df DataFrame combining each mezzanine data (Mezz, Mezz_WWPN, Mez_type)
    # in corresponding single column for all mezzanine cards
    synergy_server_wwn_df = pd.DataFrame()
    for column in mezz_columns:
        # set of columns for current mezzanine number (Mezz1, Mezz1_type, Mezz1_WWPN; next Mezz2, Mezz2_type, Mezz2_WWPN, etc)
        current_mezz_columns = [clmn for clmn in server_mezz_columns if column in clmn]
        df = synergy_mezz_df[[*server_hw_columns, *current_mezz_columns]].copy()
        df['Mezz_location'] = column
        # remove Mezz number from Mezz columns name
        mezz_rename_dct = {}
        for clmn in current_mezz_columns:
            result = re.search(pattern_dct['mezz_number_remover'], clmn)
            if not result.group(2):
                mezz_rename_dct[clmn] = result.group(1)
            else:
               mezz_rename_dct[clmn] = result.group(1) + result.group(2)
        df.rename(columns=mezz_rename_dct , inplace=True)
        if synergy_server_wwn_df.empty:
            synergy_server_wwn_df = df.copy()
        else:
            synergy_server_wwn_df = pd.concat([synergy_server_wwn_df, df], ignore_index=True)

    # extract wwpn if they exist and combine extracted wwpn in one column
    if synergy_server_wwn_df['Mezz_WWPN'].notna().any():
        synergy_server_wwn_df[['P1', 'P2']] =\
            synergy_server_wwn_df['Mezz_WWPN'].str.extract(pattern_dct['mezz_wwpn_extractor'], expand=True)
        mezz_columns = [column for column in synergy_server_wwn_df.columns if 'Mezz' in column]
        synergy_server_wwn_df.drop(columns = ['Mezz_WWPN'], inplace=True)
        combined_df = pd.DataFrame()
        for column in ['P1', 'P2']:
            synergy_server_wwn_df.rename(columns={column: 'Mezz_WWPN'}, inplace=True)
            df = synergy_server_wwn_df[[*server_hw_columns, *mezz_columns]].copy()
            synergy_server_wwn_df.drop(columns = ['Mezz_WWPN'], inplace=True)
            if combined_df.empty:
                combined_df = df.copy()
            else:
                combined_df = pd.concat([combined_df, df], ignore_index=True)

        synergy_server_wwn_df = combined_df.copy()

    synergy_server_wwn_df['device_location'] = \
        synergy_server_wwn_df[['enclosurename', 'position']].apply(dfop.wise_combine, axis=1, args=('Enclosure ', ' slot '))
    return synergy_server_wwn_df


def synergy_profile_wwn(synergy_config, synergy_server_wwn_df, pattern_dct):
    """Add mezzanine information to synergy_server_wwn_df from 'server-prof-conn-details' if sheet exist"""

    wb = load_workbook(synergy_config)
    if not 'server-prof-conn-details' in wb.sheetnames:
        return pd.DataFrame()

    syn_server_profile_connection_df = pd.read_excel(synergy_config, sheet_name='server-prof-conn-details')
    
    # server connection profile
    connection_profile_columns = ['profilename', 'portid', 'wwpn']
    mask_wwpn = syn_server_profile_connection_df['wwpn'] != 'None'
    synergy_connection_profile_df = syn_server_profile_connection_df.loc[mask_wwpn, connection_profile_columns].copy()

    synergy_connection_profile_df[['Mezz_location', 'Mezz_number']] = synergy_connection_profile_df['portid'].str.extract(pattern_dct['mezz_number_extractor'])
    synergy_connection_profile_df['Mezz_location'] = \
        synergy_connection_profile_df[['Mezz_location', 'Mezz_number']].apply(dfop.wise_combine, axis=1, args=('', ''))
    synergy_connection_profile_df.rename(columns={'profilename': 'serverprofilename', 'wwpn': 'Mezz_WWPN'}, inplace=True)
    synergy_connection_profile_df = synergy_connection_profile_df.reindex(columns=['serverprofilename', 'Mezz_location', 'Mezz_WWPN'])

    # create synergy profile details DataFrame
    synergy_profile_df = synergy_server_wwn_df.drop(columns=['Mezz_WWPN']).copy()
    synergy_profile_df.drop_duplicates(inplace=True)
    # add synerdgy server info to connection profile
    synergy_profile_wwn_df = synergy_profile_df.merge(synergy_connection_profile_df, how='left', on=['serverprofilename', 'Mezz_location'])
    return synergy_profile_wwn_df


def synergy_mezz_fw(synergy_config, synergy_servers_df, pattern_dct):
    """Add mezzanine firmware from server-fw-sw tab"""

    syn_server_fw_sw_df = pd.read_excel(synergy_config, sheet_name='server-fw-sw')
    
    # mezzaniin firmware
    server_fw_sw_columns = ['servername', 'serverprofilename', 'componentversion', 'componentlocation']
    mask_mezz = syn_server_fw_sw_df['componentlocation'].str.contains('Mezz')
    # filter Mezzanine firmware only
    synergy_mezz_fw_df = syn_server_fw_sw_df.loc[mask_mezz, server_fw_sw_columns].copy()

    synergy_mezz_fw_df['Mezz_slot'] = synergy_mezz_fw_df['componentlocation'].str.extract(pattern_dct['mezz_slot_extractor'])
    synergy_mezz_fw_df.drop(columns=['componentlocation'], inplace=True)
    synergy_servers_df['Mezz_slot'] = synergy_servers_df['Mezz_location'].str.extract(pattern_dct['mezz_slot_extractor'])

    # add fw info
    synergy_servers_df = synergy_servers_df.merge(synergy_mezz_fw_df, how='left', on=['servername', 'serverprofilename', 'Mezz_slot'])
    synergy_servers_df.drop(columns=['Mezz_slot'], inplace=True)
    return synergy_servers_df
