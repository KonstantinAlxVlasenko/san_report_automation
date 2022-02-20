"""Module to extract blade system information"""


import os
import re
import os
# import dataframe_operations as dfop
import numpy as np
import pandas as pd
from openpyxl import load_workbook


import utilities.dataframe_operations as dfop
import utilities.database_operations as dbop
import utilities.data_structure_operations as dsop
import utilities.module_execution as meop
import utilities.servicefile_operations as sfop
import utilities.filesystem_operations as fsop


def synergy_system_extract(report_entry_sr, report_creation_info_lst):
    """Function to extract blade systems information"""
    
    # report_steps_dct contains current step desciption and force and export tags
    report_constant_lst, report_steps_dct, *_ = report_creation_info_lst
    # report_constant_lst contains information: 
    # customer_name, project directory, database directory, max_title
    *_, max_title = report_constant_lst

    if pd.notna(report_entry_sr['synergy_meddler_folder']):
        synergy_folder = os.path.normpath(report_entry_sr['synergy_meddler_folder'])
    else:
        synergy_folder = None

    # names to save data obtained after current module execution
    data_names = ['synergy_interconnect', 'synergy_servers']
    # service step information
    print(f'\n\n{report_steps_dct[data_names[0]][3]}\n')

    # read data from database if they were saved on previos program execution iteration
    data_lst = dbop.read_database(report_constant_lst, report_steps_dct, *data_names)
    
    # force run when any data from data_lst was not saved (file not found) or 
    # procedure execution explicitly requested for output data or data used during fn execution  
    force_run = meop.verify_force_run(data_names, data_lst, report_steps_dct, max_title)
    if force_run:

        # lists to store only REQUIRED infromation
        # collecting data for all blades during looping
        # list containing enclosure, blade and hba information for all blade systems
        
        # list containing enclosure and interconnect modules information for all blade systems

        synergy_module_columns = ['Enclosure_Name',
                            'Enclosure_SN',
                            'Enclosure_Type',
                            'Interconnect_Bay',
                            'Interconnect_Model',
                            'Interconnect_SN',
                            'Interconnect_Firmware',
                            'Interconnect_Name',
                            'NodeName',
                            'Device_Location']

        synergy_server_columns = ['Enclosure_Name',
                                    'Enclosure_Slot',
                                    'Host_Name',
                                    'name',
                                    'serverprofilename',
                                    'Device_Model',
                                    'Device_SN',
                                    'Host_OS',
                                    'HBA_Description',
                                    'Mezz_type',
                                    'Connected_portWwn',
                                    'Mezz_location',
                                    'Device_Location',
                                    'HBA_Firmware'] 

        synergy_module_aggregated_df = pd.DataFrame(columns=synergy_module_columns)
        synergy_servers_aggregated_df = pd.DataFrame(columns=synergy_server_columns)

        if synergy_folder:    
            print('\nEXTRACTING SYNERGY SYSTEM INFORMATION ...\n')   
            
            # collects files in folder with xlsm extension
            synergy_config_lst = fsop.find_files(synergy_folder, max_title, filename_extension='xlsm')
            # number of files to check
            configs_num = len(synergy_config_lst)

            if configs_num:

                # # data imported from init file to extract values from config file
                # enclosure_params, _, comp_keys, match_keys, comp_dct = sfop.data_extract_objects('blades', max_title)
                # module_params = columns_import('blades', max_title, 'module_params')
                # blade_params = columns_import('blades', max_title, 'blade_params')

                for i, synergy_config in enumerate(synergy_config_lst):       
                    # file name with extension
                    configname_wext = os.path.basename(synergy_config)
                    # remove extension from filename
                    configname, _ = os.path.splitext(configname_wext)

                    # current operation information string
                    info = f'[{i+1} of {configs_num}]: {configname} system.'
                    print(info, end =" ")


                    wb = load_workbook(synergy_config)

                    
                    syn_enclosure_df = pd.read_excel(synergy_config, sheet_name='enclosures')
                    syn_module_df = pd.read_excel(synergy_config, sheet_name='interconnectbays')
                    
                    syn_server_hw_df = pd.read_excel(synergy_config, sheet_name='server-hardware')
                    syn_server_fw_sw_df = pd.read_excel(synergy_config, sheet_name='server-fw-sw')
                    
                    if 'server-prof-conn-details' in wb.sheetnames:
                        syn_server_profile_connection_df = pd.read_excel(synergy_config, sheet_name='server-prof-conn-details')
                    else:
                        syn_server_profile_connection_df = pd.DataFrame()

                    synergy_module_df = synergy_module(syn_enclosure_df, syn_module_df)
                    
                    if synergy_module_aggregated_df.empty:
                        synergy_module_aggregated_df = synergy_module_df
                    else:
                        synergy_module_aggregated_df = pd.concat([synergy_module_aggregated_df, synergy_module_df], ignore_index=True)
                        
                        
                    synergy_server_wwn_df = synergy_server_wwn(syn_server_hw_df)
                    
                    if not syn_server_profile_connection_df.empty:
                        synergy_profile_wwn_df = synergy_profile_wwn(syn_server_profile_connection_df, synergy_server_wwn_df)
                    else:
                        synergy_profile_wwn_df = pd.DataFrame()
                    
                    # conctenate connection profile and server hardware
                    synergy_servers_df = pd.concat([synergy_server_wwn_df, synergy_profile_wwn_df], ignore_index=True)
                    synergy_servers_df.drop_duplicates(inplace=True)
                    
                    # add mezzanine firmware details
                    synergy_servers_df = synergy_mezz_fw(syn_server_fw_sw_df, synergy_servers_df)
                    synergy_servers_df.sort_values(by=['enclosurename', 'position', 'Mezz_WWPN'], ignore_index=True, inplace=True)
                    
                    
                    if synergy_servers_aggregated_df.empty:
                        synergy_servers_aggregated_df = synergy_servers_df
                    else:
                        synergy_servers_aggregated_df = pd.concat([synergy_servers_aggregated_df, synergy_servers_df], ignore_index=True)
                        
                    if synergy_module_aggregated_df['switchbasewwn'].notna().any():
                        synergy_module_aggregated_df['switchbasewwn'] = synergy_module_aggregated_df['switchbasewwn'].str.lower()
                    if synergy_servers_aggregated_df['Mezz_WWPN'].notna().any():
                        synergy_servers_aggregated_df['Mezz_WWPN'] = synergy_servers_aggregated_df['Mezz_WWPN'].str.lower()


                    if not synergy_servers_df.empty or not synergy_module_df.empty:
                        meop.status_info('ok', max_title, len(info))
                    else:
                        meop.status_info('no data', max_title, len(info))

                module_columns_dct = {'enclosurename': 'Enclosure_Name',
                                    'enclosure_serialnumber': 'Enclosure_SN',
                                    'enclosuretype': 'Enclosure_Type',
                                    'baynumber': 'Interconnect_Bay',
                                    'interconnectmodel': 'Interconnect_Model',
                                    'serialnumber': 'Interconnect_SN',
                                    'switchfwversion': 'Interconnect_Firmware',
                                    'hostname': 'Interconnect_Name',
                                    'switchbasewwn': 'NodeName',
                                    'device_location': 'Device_Location'}

                synergy_module_aggregated_df.rename(columns=module_columns_dct, inplace=True)
                synergy_module_aggregated_df.replace(r'^None$|^none$|^ *$', value=np.nan, regex=True, inplace=True)

                server_columns_dct = {'enclosurename': 'Enclosure_Name',
                                    'position': 'Enclosure_Slot',
                                    'servername': 'Host_Name',
                                    'model': 'Device_Model',
                                    'serialnumber': 'Device_SN',
                                    'oshint': 'Host_OS',
                                    'Mezz': 'HBA_Description',
                                    'Mezz_WWPN': 'Connected_portWwn',
                                    'device_location': 'Device_Location',
                                    'componentversion': 'HBA_Firmware'} 

                synergy_servers_aggregated_df.rename(columns=server_columns_dct, inplace=True)
                synergy_servers_aggregated_df.replace(r'^None$|^none$|^ *$', value=np.nan, regex=True, inplace=True)

                # data_lst = [synergy_module_aggregated_df, synergy_servers_aggregated_df]
                # # save extracted data to json file
                # save_data(report_constant_lst, data_names, *data_lst)
        else:
            # current operation information string
            info = f'Collecting synergy details'
            print(info, end =" ")
            meop.status_info('skip', max_title, len(info))
        data_lst = [synergy_module_aggregated_df, synergy_servers_aggregated_df]
        # write data to sql db
        dbop.write_database(report_constant_lst, report_steps_dct, data_names, *data_lst)  
    # verify if loaded data is empty after first iteration and replace information string with empty list
    else:
        data_lst = dbop.verify_read_data(report_constant_lst, data_names, *data_lst)
        synergy_module_aggregated_df, synergy_servers_aggregated_df = data_lst

    # save data to service file if it's required
    for data_name, data_frame in zip(data_names, data_lst):
        dfop.dataframe_to_excel(data_frame, data_name, report_creation_info_lst)
    return synergy_module_aggregated_df, synergy_servers_aggregated_df


def synergy_module(syn_enclosure_df, syn_module_df):
    
    # enclosure information
    enclosure_columns = ['name', 'enclosuremodel', 'version', 'serialnumber']
    synergy_enclosure_df = syn_enclosure_df[enclosure_columns].copy()
    
    synergy_enclosure_df['enclosuretype'] = \
        synergy_enclosure_df[['enclosuremodel', 'version']].apply(dfop.wise_combine, axis=1)
        
    synergy_enclosure_df.rename(columns={'name': 'enclosurename', 'serialnumber': 'enclosure_serialnumber'}, inplace=True)
    synergy_enclosure_df.drop(columns=['enclosuremodel', 'version'], inplace=True)
    
    # interconnect information
    module_columns = ['enclosurename',  'baynumber',  'interconnectmodel',  'serialnumber',  'switchfwversion',  'hostname',  'switchbasewwn']
    mask_state = syn_module_df['state'].notna()
    synergy_module_df = syn_module_df.loc[mask_state, module_columns]
    synergy_module_df['device_location'] = \
        synergy_module_df[['enclosurename', 'baynumber']].apply(dfop.wise_combine, axis=1, args=('Enclosure ', ' bay '))
    
    # add interconnect information to enclosure
    synergy_module_df = synergy_enclosure_df.merge(synergy_module_df, how='left', on=['enclosurename'])
    
    return synergy_module_df
            

def synergy_server_wwn(syn_server_hw_df):

    # server_hardware
    server_hw_columns = [ 'enclosurename', 'position', 'servername',  'name',  'serverprofilename',  'model',  'serialnumber',  'oshint']
    # columns with Mezz in name except MAC columns
    server_mezz_columns = [column for column in syn_server_hw_df.columns if 
                           re.match(r'^Mezz\d+(?!_MAC)\w*$', column)]
    
    server_hw_mezz_columns = [*server_hw_columns, *server_mezz_columns]
    # filter servers (filter off composer) 
    mask_profile = syn_server_hw_df['serverprofilename'] != 'None'    
    # create DataFrame with server mezzanin cards
    synergy_mezz_df = syn_server_hw_df.loc[mask_profile, server_hw_mezz_columns].copy()
    


    # filter off empty mezzanin slots and columns except Mezz
    mezz_columns = [column for column in server_mezz_columns if 
                    re.match(r'Mezz\d$', column) and 
                    (synergy_mezz_df[column] != '<empty>').all()]
    
    
    # create synergy_server_wwn_df DataFrame combining each mezzanine data (Mezz, Mezz_WWPN, Mez_type) 
    # in corresponding single column for all mezzanine cards
    synergy_server_wwn_df = pd.DataFrame()
    for column in mezz_columns:
        # set of columns for current mezzanine number (Mezz1, Mezz1_type, Mezz1_WWPN; next Mezz2, Mezz2_type, Mezz2_WWPN, etc)
        current_mezz_columns = [clmn for clmn in server_mezz_columns if column in clmn]
        df = synergy_mezz_df[[*server_hw_columns, *current_mezz_columns]].copy()
        df['Mezz_location'] = column
        # remove Mezz number from Mezz columns name
        mezz_rename_dct = {}
        for clmn in current_mezz_columns:
            result = re.search(r'^(Mezz)\d(.*)$', clmn)
            if not result.group(2):
                mezz_rename_dct[clmn] = result.group(1)
            else:
               mezz_rename_dct[clmn] = result.group(1) + result.group(2)
        df.rename(columns=mezz_rename_dct , inplace=True)                                            
        if synergy_server_wwn_df.empty:
            synergy_server_wwn_df = df.copy()
        else:
            synergy_server_wwn_df = pd.concat([synergy_server_wwn_df, df], ignore_index=True)
            
    # extract wwpn if they exist and combine extracted wwpn in one column
    if synergy_server_wwn_df['Mezz_WWPN'].notna().any():
        synergy_server_wwn_df[['P1', 'P2']] =\
            synergy_server_wwn_df['Mezz_WWPN'].str.extract(r' *P\d=((?:[0-9a-fA-F]{2}:){7}[0-9a-fA-F]{2}) +P\d=((?:[0-9a-fA-F]{2}:){7}[0-9a-fA-F]{2})')
        mezz_columns = [column for column in synergy_server_wwn_df.columns if 'Mezz' in column]
        synergy_server_wwn_df.drop(columns = ['Mezz_WWPN'], inplace=True)
        combined_df = pd.DataFrame()
        for column in ['P1', 'P2']:
            synergy_server_wwn_df.rename(columns={column: 'Mezz_WWPN'}, inplace=True)
            df = synergy_server_wwn_df[[*server_hw_columns, *mezz_columns]].copy()
            synergy_server_wwn_df.drop(columns = ['Mezz_WWPN'], inplace=True)
            if combined_df.empty:
                combined_df = df.copy()
            else:
                combined_df = pd.concat([combined_df, df], ignore_index=True)
            
        synergy_server_wwn_df = combined_df.copy()
            
    synergy_server_wwn_df['device_location'] = \
        synergy_server_wwn_df[['enclosurename', 'position']].apply(dfop.wise_combine, axis=1, args=('Enclosure ', ' slot '))

    return synergy_server_wwn_df
    


def synergy_profile_wwn(syn_server_profile_connection_df, synergy_server_wwn_df):
    
    # server connection profile
    connection_profile_columns = ['profilename', 'portid', 'wwpn']
    mask_wwpn = syn_server_profile_connection_df['wwpn'] != 'None'
    synergy_connection_profile_df = syn_server_profile_connection_df.loc[mask_wwpn, connection_profile_columns].copy()
    
    synergy_connection_profile_df[['Mezz_location', 'Mezz_number']] = synergy_connection_profile_df['portid'].str.extract(r'^ *(\w+) *(\d)')
    synergy_connection_profile_df['Mezz_location'] = \
        synergy_connection_profile_df[['Mezz_location', 'Mezz_number']].apply(dfop.wise_combine, axis=1, args=('', ''))
    synergy_connection_profile_df.rename(columns={'profilename': 'serverprofilename', 'wwpn': 'Mezz_WWPN'}, inplace=True)
    synergy_connection_profile_df = synergy_connection_profile_df.reindex(columns=['serverprofilename', 'Mezz_location', 'Mezz_WWPN'])
    
    
    # create synergy profile details DataFrame
    synergy_profile_df = synergy_server_wwn_df.drop(columns=['Mezz_WWPN']).copy()
    synergy_profile_df.drop_duplicates(inplace=True)
    
    # add synerdgy server info to connection profile
    synergy_profile_wwn_df = synergy_profile_df.merge(synergy_connection_profile_df, how='left', on=['serverprofilename', 'Mezz_location'])
    
    return synergy_profile_wwn_df
    
    
    
def synergy_mezz_fw(syn_server_fw_sw_df, synergy_servers_df):

    # mezzaniin firmware
    server_fw_sw_columns = ['servername', 'serverprofilename', 'componentversion', 'componentlocation']
    mask_mezz = syn_server_fw_sw_df['componentlocation'].str.contains('Mezz')
    # filter Mezzanine firmware only
    synergy_mezz_fw_df = syn_server_fw_sw_df.loc[mask_mezz, server_fw_sw_columns].copy()
    
    synergy_mezz_fw_df['Mezz_slot'] = synergy_mezz_fw_df['componentlocation'].str.extract(r'.(\d+)')
    synergy_mezz_fw_df.drop(columns=['componentlocation'], inplace=True)
    synergy_servers_df['Mezz_slot'] = synergy_servers_df['Mezz_location'].str.extract(r'.(\d+)')
    
    # add fw info
    synergy_servers_df = synergy_servers_df.merge(synergy_mezz_fw_df, how='left', on=['servername', 'serverprofilename', 'Mezz_slot'])
    synergy_servers_df.drop(columns=['Mezz_slot'], inplace=True)
    
    return synergy_servers_df
