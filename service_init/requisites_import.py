"""Moodule to import service prerequisites, poroject steps force keys and description, visio templates path and device rack location paths"""

import os
import sys

import pandas as pd
from openpyxl import load_workbook

import utilities.dataframe_operations as dfop
import utilities.filesystem_operations as fsop
import utilities.module_execution as meop
import utilities.servicefile_operations as sfop


def import_requisites(max_title):
    """Function to import entry report values:
    customer_name, project_title, hardware configuration files folders"""

    report_requisites_df = sfop.dataframe_import('report_requisites', max_title, 'report_info.xlsx', display_status=False)
    report_requisites_sr = dfop.series_from_dataframe(report_requisites_df, index_column='name', value_column='value')

    # fields which shouldn't be empty
    entry_lst = ('customer_name', 'project_title', 'project_folder', 'supportsave_folder')
    empty_entry_lst = [entry for entry in  entry_lst if pd.isna(report_requisites_sr[entry])]

    if empty_entry_lst:
        print(f"{', '.join(empty_entry_lst)} {'are' if len(empty_entry_lst)>1 else 'is'} not defined in the report_info file.")
        sys.exit()
    
    report_requisites_sr['project_title'] = report_requisites_sr['project_title'].replace(' ', '_')
    report_requisites_sr['customer_name'] = report_requisites_sr['customer_name'].replace(' ', '_')
    # check if directories and files in requisites are valid
    validate_requisites_paths(report_requisites_sr)
    return report_requisites_sr


def validate_requisites_paths(report_requisites_sr):
    """Function to check if all directories and file paths defind in requsites  exist.
    If any directory doesn't exist program stops exection.
    If file paths doesn't exist request for continue program exection is shown"""

    for index, _ in report_requisites_sr.items():
        # check defined folders and file paths only
        if pd.isna(report_requisites_sr[index]):
            report_requisites_sr[index] = None
            continue
        if not any(['folder' in index, 'path' in index]):
            continue
        
        # normpath all folders and file paths
        report_requisites_sr[index] = os.path.normpath(report_requisites_sr[index])
        # check if folder exist
        if 'folder' in index:
            fsop.check_valid_path(report_requisites_sr[index])
        # check if file exist
        elif 'path' in index:
            if not fsop.validate_path_isfile(report_requisites_sr[index]):
                print(f"\nFile '{report_requisites_sr[index]}' is missiing and will be removed from the requisites.")
                meop.continue_request()
                report_requisites_sr[index] = None


def validate_device_rack_file(device_rack_file, max_title):
    """Function validates if device rack file exist and has valid format.
    If not request for program execution is shown"""

    if pd.notna(device_rack_file):
        # check if file excel file
        if not fsop.validate_excel_file(device_rack_file):
            print(f"\nFile '{os.path.basename(device_rack_file)}'", 
                    "is not excel file, and will be removed from the requisites.")
            meop.continue_request()
            device_rack_file = None
        # check if file has 'switch_rack' tab
        else:
            wb = load_workbook(device_rack_file, read_only=True)
            if not 'switch_rack' in wb.sheetnames:
                print(f"\nFile '{os.path.basename(device_rack_file)}'",
                        "has no 'switch_rack' sheet and will be removed from the requisites.")
                meop.continue_request()
                device_rack_file = None

    # check if file has 'switchWwn', 'Device_Rack' columns
    if pd.notna(device_rack_file):
        switch_rack_df = sfop.dataframe_import('switch_rack', max_title, 
                                                init_file=device_rack_file, 
                                                header=0, display_status=False)
        if not dfop.verify_columns_in_dataframe(switch_rack_df, columns=['switchWwn', 'Device_Rack']):
            absent_columns_lst = [column for column in ['switchWwn', 'Device_Rack'] if not column in switch_rack_df.columns]
            print(f"\nSwitch rack details dataframe missing '{', '.join(absent_columns_lst)}' column(s)", 
                    "and will be removed from the requisites.")
            meop.continue_request()
            device_rack_file = None
    return device_rack_file


def import_service_dataframes(max_title):
    """Function to import tables related information: table name, type, description, order in report,
    headers, correlation, force_extract and export to excel keys."""

    print(f'\n\nPREREQUISITES 2. IMPORTING STEPS AND TABLES RELATED INFORMATION\n')

    project_steps_df = import_project_steps(max_title)

    # DataFrame with report column names
    report_headers_df = sfop.dataframe_import('customer_report', max_title)
    # software path
    software_path_df = sfop.dataframe_import('software', max_title, init_file = 'report_info.xlsx')
    software_path_df['path'] = software_path_df.apply(lambda series: software_path(series), axis=1)
    software_path_sr = dfop.series_from_dataframe(software_path_df, index_column='name', value_column='path')
    # DataFrame with input and output data names of each module
    io_data_names_df = sfop.dataframe_import('in_out_data_names', max_title, init_file='report_info.xlsx')
    # constants to create SAN topology in Visio
    san_topology_constantants_df = sfop.dataframe_import('san_topology_constants', max_title, init_file='report_info.xlsx')
    san_topology_constantants_sr = dfop.series_from_dataframe(san_topology_constantants_df, index_column='name', value_column='value')
    # import data with switch models, firmware and etc
    san_graph_grid_df = sfop.dataframe_import('san_graph_grid', max_title)
    return project_steps_df, io_data_names_df, report_headers_df, software_path_sr, san_graph_grid_df, san_topology_constantants_sr


def import_project_steps(max_title):
    """Function to import project steps details (step description, 'export_to_excel', 'force_run' keys 
    and sort weights for sorting sheets in report) """

    project_steps_df = sfop.dataframe_import('project_steps', max_title, init_file='report_info.xlsx')
    
    # replace all nan values with 0 and all notna values (except 0) with 1
    project_steps_df.fillna({'export_to_excel': 0, 'force_run': 0, 'sort_weight': 1, 
                            'report_type': 'unknown', 'module_info': '-', 'step_info': '-', 'description': '-', }, inplace=True)
    for column in ['export_to_excel', 'force_run']: 
        mask_force_run = project_steps_df[column].notna() & ~project_steps_df[column].isin([0, '0'])
        project_steps_df.loc[mask_force_run, column] = 1

    numeric_columns = ['export_to_excel', 'force_run', 'sort_weight']
    project_steps_df[numeric_columns] = project_steps_df[numeric_columns].apply(pd.to_numeric, errors='ignore')

    info = "Global export report key"
    print(info, end =" ") 

    # check if global export report key is set
    mask_global_report = project_steps_df['keys'] == 'report'
    if project_steps_df.loc[mask_global_report, 'export_to_excel'].values != 0:
        mask_report_type = project_steps_df['report_type'] == 'report'
        project_steps_df.loc[mask_report_type, 'export_to_excel'] = 1
        meop.status_info('on', max_title, len(info))
    else:
        meop.status_info('off', max_title, len(info))
    
    project_steps_columns =  ['keys', 'report_type', 'export_to_excel', 'force_run', 
                                'module_info', 'step_info', 'description', 'sort_weight']
    project_steps_df = project_steps_df[project_steps_columns].copy()
    project_steps_df.set_index('keys', inplace=True)
    return project_steps_df


def software_path(series):
    """Function to find software pathes for san_toolbox and s3mft"""
    
    sw_path = os.path.join(series['directory'], series['file'])
    sw_path = os.path.normpath(sw_path)
    return sw_path