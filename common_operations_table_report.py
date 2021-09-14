
import os
import sys
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from common_operations_miscellaneous import status_info


# saving DataFrame to excel file
def dataframe_to_report(data_frame, sheet_title, report_creation_info_lst, current_date=str(date.today()), force_flag = False, freeze_column='A'):
    """Check if excel file exists, write DataFrame, create or update table of contents,
    change DataFrame text and presentation format."""
    
    report_constant_lst, report_steps_dct, *_ = report_creation_info_lst
    customer_name, report_path, _, max_title = report_constant_lst
    
    # report_steps_dct for each data_name contains: export_to_excel flag, 
    # force_extract flag, report_type, step_info, data_description
    if report_steps_dct.get(sheet_title):
        export_flag, _, report_type, _, df_decription = report_steps_dct.get(sheet_title)
    else:
        info = f'DataFrame {sheet_title}'
        print(info, end =" ")
        status_info('unknown', max_title, len(info))
        export_flag, report_type = 1, 'unknown'
        df_decription = '-'

    # check DataFrame report type to save
    if report_type == 'report':
        report_mark = 'SAN_Assessment_Tables'
    else:
        report_mark = report_type
          
    # construct excel filename
    file_name = customer_name + '_' + report_mark + '_' + current_date + '.xlsx'

    # information string
    info = f'Exporting {sheet_title} table to {report_mark} file'
    print(info, end =" ")
    file_path = os.path.join(report_path, file_name)
    
    # save DataFrame to excel file if export_to_excel trigger is ON
    # and DataFrame is not empty
    if (force_flag or export_flag) and not data_frame.empty:
        file_mode = 'a' if os.path.isfile(file_path) else 'w'
        try:
            if_sheet_exists_param = 'replace' if file_mode == 'a' else None
            with pd.ExcelWriter(file_path, mode=file_mode, if_sheet_exists=if_sheet_exists_param) as writer:  # pylint: disable=abstract-class-instantiated
                table_of_contents_generation(writer, file_mode, sheet_title, df_decription)
                write_dataframe_to_worksheet(writer, data_frame, sheet_title)
                workbook = openpyxl.load_workbook(writer)
                writer.book = workbook
                writer.sheets = {ws.title: ws for ws in workbook.worksheets}
                # add DataFrame title and link to the table of contents
                add_dataframe_title(workbook, sheet_title, df_decription)
                # change text format in header (wraped, bold, size 10) and data rows (size 10)
                worksheet_format(workbook, sheet_title, freeze_column, header_row_num=3, font_size=10)
                writer.save()
        except PermissionError:
            status_info('fail', max_title, len(info))
            print('\nPermission denied. Close the file.\n')
            sys.exit()
        else:
            status_info('ok', max_title, len(info))
        return file_path        
    else:
        # if save key is on but DataFrame empty
        if report_steps_dct[sheet_title][0] and data_frame.empty:
            status_info('no data', max_title, len(info))
        else:            
            status_info('skip', max_title, len(info))
        return None


def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet):
        """Make all columns best fit"""

        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].bestFit = True


def add_dataframe_title(workbook, sheet_title, df_decription):
    """Function to add DataFrame description and link to the table of contents
    to the first two rows of the worksheet"""

    # add table title
    workbook[sheet_title]['A1'] = df_decription
    workbook[sheet_title]['A1'].font =  Font(bold=True)
    # add hyperlink to the Table of contents
    create_hyperlink(workbook[sheet_title], 'A2', 'Содержание', cell_ref='A2', display_name='К содержанию')
                

def worksheet_format(workbook, sheet_title: str, freeze_column: str, header_row_num: int, font_size: int, font_name=None):
    """Function to change worksheet text format.
    Header is bold and wrapText oprion is used. Header and data font_size changed.
    Cells in columns are best fit"""                
    
    # change font, wrap_text header and fit text cell size
    ws = workbook[sheet_title]
    row_number = 1
    for row in ws.rows:
        for cell in row:
            if row_number == header_row_num:
                cell.font = Font(size=font_size, bold=True)
                cell.alignment = Alignment(wrapText=True)
            if row_number > header_row_num:
                cell.font = Font(size=font_size)
        row_number += 1 
    columns_best_fit(ws)
    # for idx, col in enumerate(ws.columns, 1):
    #     ws.column_dimensions[get_column_letter(idx)].auto_size = True

    # freeze header
    freeze_cell = freeze_column + str(header_row_num + 1)
    ws.freeze_panes = freeze_cell

def write_dataframe_to_worksheet(writer, df, sheet_title):
    """Auxiliary function to write DataFrane to excel.
    If DataFrame index is MultiIndex then index is reset."""

    # reset index if DataFrame have MultiIndex
    if isinstance(df.index, pd.MultiIndex):
        df_flat = df.reset_index()
    # keep index if False
    else:
        df_flat = df.copy()
    # writing DataFrame with single Index
    df_flat.to_excel(writer, sheet_name=sheet_title,  startrow=2, index=False)
    writer.save()


def table_of_contents_generation(writer, file_mode, sheet_title, df_decription):
    """"Function to create or recreate table of contents if file is newly created or new sheet added"""

    item_exist = True 

    # if file exists and there is no sheet_title in contents add new item to the existing content
    if  file_mode == 'a':
        existing_content_df = pd.read_excel(writer, sheet_name='Содержание')
        item_exist = sheet_title in existing_content_df['Закладка'].values
        if not item_exist:
            # content item to add
            content_df = pd.DataFrame([[sheet_title, df_decription]], columns=['Закладка', 'Название таблицы'])
            # concatenate existing items to the current
            content_df = pd.concat([existing_content_df, content_df])
    # if file is newly created then create contents with single sheet_title
    elif file_mode == 'w':
        content_df = pd.DataFrame([[sheet_title, df_decription]], columns=['Закладка', 'Название таблицы'])

    # save contents and create hyperlinks to the corresponding sheets
    if file_mode == 'w' or not item_exist:
        # write content to excel file
        content_df.to_excel(writer, sheet_name='Содержание', index = False)
        writer.save()
         
        workbook = openpyxl.load_workbook(writer)
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets) 

        ws_content = workbook['Содержание']
        # create hyperlinks for all items except header
        row_number = 1
        for row in ws_content.values:
            # skip header
            if row_number > 1:
                bookmark_title = row[0]
                bookmark_addr = 'A' + str(row_number)
                create_hyperlink(ws_content, bookmark_addr, sheet_name=bookmark_title, cell_ref='A1', display_name=None)
            row_number += 1
        # change cell size for best fit
        columns_best_fit(ws_content)
        # for idx, col in enumerate(ws_content.columns, 1):
        #     ws_content.column_dimensions[get_column_letter(idx)].auto_size = True
        
        # freeze header
        ws_content.freeze_panes = 'B2'
        writer.save()


def create_hyperlink(ws, at_cell, sheet_name, cell_ref='A1', display_name=None):
    """Function to create hyperlinks in the worksheet ws to the sheet_name"""

    if display_name is None:
        display_name = sheet_name
    to_location = "'{0}'!{1}".format(sheet_name, cell_ref)
    ws[at_cell].hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(display=display_name, ref=at_cell, location=to_location)
    ws[at_cell].value = display_name
    ws[at_cell].font =  openpyxl.styles.fonts.Font(u='single', color=openpyxl.styles.colors.BLUE)    
