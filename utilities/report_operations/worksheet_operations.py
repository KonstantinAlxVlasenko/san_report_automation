"""Module to format worksheet in report excel file"""

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def format_workbook(workbook, sheet_title, df_decription, freeze_column):
    """Function to format table of contents and data worksheets"""

    # create hyperlinks for all items of table of contents
    hyperlink_content(workbook)
    # add DataFrame description and link to the table of contents
    add_dataframe_title(workbook, sheet_title, df_decription)
    # change worksheet text format
    format_worksheet(workbook, sheet_title, font_size=10)
    # freeze header row
    freeze_header(workbook, sheet_title, freeze_column)
    # add filter functionality to the header row
    add_header_filter(workbook, sheet_title)


def hyperlink_content(workbook):
    """Function to create hyperlinks for all items of table of contents"""

    ws_content = workbook['Содержание']
    # create hyperlinks for all items except header
    row_number = 1
    for row in ws_content.values:
        # skip header
        if row_number > 1:
            bookmark_title = row[0]
            bookmark_addr = 'A' + str(row_number)
            create_hyperlink(ws_content, bookmark_addr, sheet_name=bookmark_title, cell_ref='A1', display_name=None)
        row_number += 1
    # change cell size for best fit
    columns_best_fit(ws_content)    
    # freeze header
    ws_content.freeze_panes = 'B2'


def create_hyperlink(ws, at_cell, sheet_name, cell_ref='A1', display_name=None):
    """Function to create hyperlinks in the worksheet ws to the sheet_name"""

    if display_name is None:
        display_name = sheet_name
    to_location = "'{0}'!{1}".format(sheet_name, cell_ref)
    ws[at_cell].hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(display=display_name, ref=at_cell, location=to_location)
    ws[at_cell].value = display_name
    ws[at_cell].font =  openpyxl.styles.fonts.Font(u='single', color=openpyxl.styles.colors.BLUE)


def add_dataframe_title(workbook, sheet_title, df_decription):
    """Function to add DataFrame description and link to the table of contents
    to the first two rows of the worksheet"""

    # add table title
    workbook[sheet_title]['A1'] = df_decription
    workbook[sheet_title]['A1'].font =  Font(bold=True)
    # add hyperlink to the Table of contents
    create_hyperlink(workbook[sheet_title], at_cell='A2', sheet_name='Содержание', cell_ref='A2', display_name='К содержанию')
    

def format_worksheet(workbook, sheet_title: str, font_size: int, font_name=None, header_row_num: int=3):
    """Function to change worksheet text format.
    Header is bold and wrapText option is used (to avoid truncation). Header and data font_size changed.
    Cells in columns are best fit"""                
    
    # change font, wrap_text header and fit text cell size
    ws = workbook[sheet_title]
    row_number = 1
    for row in ws.rows:
        for cell in row:
            if row_number == header_row_num:
                cell.font = Font(size=font_size, bold=True)
                cell.alignment = Alignment(wrapText=True)
            if row_number > header_row_num:
                cell.font = Font(size=font_size)
        row_number += 1 
    columns_best_fit(ws)
    

def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet):
        """Make all columns best fit"""

        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].bestFit = True


def freeze_header(workbook, sheet_title: str, freeze_column: str, header_row_num: int=3):
    """Function to freeze header row"""

    worksheet = workbook[sheet_title]
    freeze_cell = freeze_column + str(header_row_num + 1)
    worksheet.freeze_panes = freeze_cell
    

def add_header_filter(workbook, sheet_title, header_row_num: int=3):
    """Function to add filter functionality to the header row"""

    worksheet = workbook[sheet_title]
    FullRange = "A" + str(header_row_num) + ":" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
